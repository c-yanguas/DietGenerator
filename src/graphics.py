import matplotlib.pyplot as plt
import random
import pandas as pd
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.drawing.image import Image
import seaborn as sns 
import os
import constants as ct
import locale

idioma_sistema = locale.getdefaultlocale()[0]
if idioma_sistema:
    sheet_name = "Hoja 1" if idioma_sistema.startswith("es") else "Sheet 1"
else:
    sheet_name = "Hoja 1"

def generar_grafico(datos, nombre_archivo=ct.IMAGES_PATH_TMP):
    plt.figure(figsize=(6, 4))
    # Colores para los primeros cinco elementos
    colores_principales = ['lightcoral', 'lightblue', 'lightgreen', 'orange', 'lightyellow']

    # Colores adicionales para elementos adicionales
    colores_adicionales = [(random.random(), random.random(), random.random(), 1.0) for _ in range(len(datos) - 5)]

    # Combinar los colores
    colores = colores_principales + colores_adicionales

    # Crear un gráfico de pastel
    plt.pie(datos.values(), labels=datos.keys(), autopct='%1.1f%%', colors=colores, startangle=90)

    # Agregar leyenda
    # plt.legend()
    plt.legend(labels=datos.keys(), loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=len(datos)//2, fancybox=True, shadow=True)


    # Agregar titulo
    plt.title('Distribución de Macronutrientes', fontsize=16)


    # Ajustar el aspecto para que el gráfico se vea como un círculo
    plt.axis('equal')

    # Guardar el gráfico como imagen
    plt.savefig(nombre_archivo)

    plt.close()

    return nombre_archivo


def generar_grafico_opcion_2(datos, nombre_archivo=ct.IMAGES_PATH_TMP):
    if os.path.exists(nombre_archivo):
        os.remove(nombre_archivo)
    # Paleta de colores profesional de seaborn
    colores = sns.color_palette('Paired', len(datos))

    # Crear un gráfico de pastel
    plt.figure(figsize=(4, 6))
    plt.pie(datos.values(), labels=datos.keys(), autopct='%1.1f%%', colors=colores, startangle=90, wedgeprops=dict(width=0.4, edgecolor='w'))

    # Añadir leyenda y título
    plt.legend(labels=datos.keys(), loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=len(datos)//2, fancybox=True, shadow=True)

    # Agregar título
    plt.title('Distribución de Macronutrientes', fontsize=16)

    # Ajustar el aspecto para que el gráfico se vea como un círculo
    plt.axis('equal')

    # Agregar un marco negro ligero alrededor del gráfico
    plt.gca().add_artist(plt.Circle((0,0),0.7,fc='black'))

    # Guardar el gráfico como imagen con un marco blanco alrededor
    plt.savefig(nombre_archivo, bbox_inches='tight', pad_inches=0.1)

    plt.clf()

    return nombre_archivo



def pintar_celdas_por_criterios(df, criterios, nombre_archivo):
    """
    Se pintan las celdas del excel en base a un conjunto de criterios.
    """
    # Crear un escritor de Excel
    excel_writer = pd.ExcelWriter(nombre_archivo, engine='openpyxl')

    # Volcar el DataFrame a Excel
    df.to_excel(excel_writer, sheet_name=sheet_name, index=True)

    # Obtener la hoja de Excel
    excel_sheet = excel_writer.sheets[sheet_name]

    # Iterar a través de la lista de criterios y resaltar las celdas según cada criterio
    for criterio in criterios:
        color = criterio['color']
        columna_filtro = criterio['columna_filtro']
        valor_filtro = criterio['valor_filtro']

        # Crear un formato de relleno para resaltar las celdas
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Filtrar las celdas a pintar
        filtered_df_index = df.loc[df.index.get_level_values(columna_filtro) == valor_filtro].index

        # Definir el numero de columnas a colorear, es decir, los indices + columnas planas. + 1 por el range de abajo
        index_to_color = len(df.index[0])
        cols_to_color = len(df.columns)
        cols_to_color = index_to_color + cols_to_color + 1

        # Acceder a las celdas del DataFrame original y resaltarlas
        for index in filtered_df_index:
            # + 2, bc df idx start at 0, excel start at 1 and 1 is header
            num_row_to_color = df.index.get_loc(index) + 2
            # range 2, para evitar colorear la primera col 
            for col_num in range(2, cols_to_color):
                cell = excel_sheet.cell(row=num_row_to_color, column=col_num)
                cell.fill = fill

    # Ajustar automáticamente el ancho de las columnas al contenido
    for column in excel_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        excel_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Guardar el archivo Excel nuevamente después de ajustar el ancho de las columnas
    excel_writer.close()


    return nombre_archivo




def insertar_imagen_en_excel(excel_path, imagen_path, nombre_hoja='Sheet1', celda='K1'):
    # Cargar el libro de trabajo de Excel
    libro_excel = openpyxl.load_workbook(excel_path)

    # Obtener la hoja de trabajo (puedes cambiar 'Sheet1' por el nombre de la hoja deseada)
    nombre_hoja = sheet_name
    hoja_trabajo = libro_excel[nombre_hoja]

    # Crear un objeto Image a partir del archivo de imagen
    imagen = Image(imagen_path)

    # Establecer la posición de la celda en la que deseas insertar la imagen
    imagen.anchor = celda

    # Insertar la imagen en la hoja de trabajo
    hoja_trabajo.add_image(imagen)

    # Guardar los cambios en el libro de trabajo
    libro_excel.save(excel_path)



def graficar_evolucion_dieta(path_historic: str, nombre_archivo=ct.IMAGES_PATH_TMP):
    # Leer los datos históricos del cliente
    df = pd.read_excel(path_historic, dtype={ct.ClientHistory.Columns.DATE: str})
    columns_to_plot = ct.get_attr_values_list_from_class(ct.ClientHistory.ColumnsMandatoryToPlot)
    df = df[columns_to_plot]

    # Convertir la columna 'Fecha' a tipo datetime
    df[ct.ClientHistory.Columns.DATE] = pd.to_datetime(df[ct.ClientHistory.Columns.DATE], format=ct.ClientHistory.DATE_FORMAT)

    # Obtener la lista de columnas a graficar
    columnas = ct.get_attr_values_list_from_class(ct.ClientHistory.ColumnsToPlot)

    # Obtener el número de columnas
    num_columnas = len(columnas)
    graphs_per_row = 3

    # Calcular el número de filas y columnas necesario
    num_filas = (num_columnas + graphs_per_row - 1) // graphs_per_row
    num_cols_last_row = num_columnas % graphs_per_row if num_columnas % graphs_per_row != 0 else graphs_per_row

    # Crear subplots con el número adecuado de filas y columnas
    fig, axs = plt.subplots(num_filas, graphs_per_row, figsize=(12, 3 * num_filas), sharex=False)
    axs = axs.flatten()

    # Iterar sobre las columnas y crear un subplot para cada una
    for i, columna in enumerate(columnas):
        ax = axs[i]
        ax.plot(df[ct.ClientHistory.Columns.DATE], df[columna], label=columna)

        # Encontrar el índice del máximo y mínimo de la columna actual
        idx_max = df[columna].idxmax()
        idx_min = df[columna].idxmin()

        max_value = df[columna][idx_max]
        min_value = df[columna][idx_min]

        # Marcar el máximo y mínimo con puntos del mismo color que la línea
        ax.plot(df[ct.ClientHistory.Columns.DATE][idx_max], max_value, marker='o', markersize=8, color='red', label='Max')
        ax.plot(df[ct.ClientHistory.Columns.DATE][idx_min], min_value, marker='o', markersize=8, color='blue', label='Min')

        ax.set_title(columna)
        ax.legend()

        # Ajustar la rotación de las etiquetas del eje X
        ax.tick_params(axis='x', rotation=45)

    # Eliminar gráficos vacíos si no hay suficientes columnas para llenar la última fila
    for j in range(num_columnas, num_filas * graphs_per_row):
        fig.delaxes(axs[j])

    # Ajustes generales
    plt.xlabel(ct.ClientHistory.Columns.DATE)
    plt.suptitle(f"Evolución de dieta de {df[ct.ClientConfiguration.Column.CLIENT].values[0]}", y=1.02)
    plt.tight_layout()

    # Guardar el gráfico como imagen con un marco blanco alrededor
    plt.savefig(nombre_archivo, bbox_inches='tight', pad_inches=0.1)

    plt.close()

    return nombre_archivo