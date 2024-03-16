import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.drawing.image import Image
import os
import shutil
from copy import copy

def pintar_celdas_por_criterios(df, criterios, nombre_archivo):
    """
    Se pintan las celdas del excel en base a un conjunto de criterios.
    """
    # Crear un escritor de Excel
    excel_writer = pd.ExcelWriter(nombre_archivo, engine='openpyxl')

    # Volcar el DataFrame a Excel
    df.to_excel(excel_writer, sheet_name='Sheet1', index=True)

    # Obtener la hoja de Excel
    excel_sheet = excel_writer.sheets['Sheet1']

    # Iterar a través de la lista de criterios y resaltar las celdas según cada criterio
    for criterio in criterios:
        color = criterio['color']
        columna_filtro = criterio['columna_filtro']
        valor_filtro = criterio['valor_filtro']

        # Crear un formato de relleno para resaltar las celdas
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Filtrar las celdas a pintar
        filtered_df_index = df.loc[df.index.get_level_values(columna_filtro) == valor_filtro].index

        # Definir el numero de columnas a colorear, es decir, los indices + columnas planas. + 1 por el range de abajo
        index_to_color = len(df.index[0])
        cols_to_color = len(df.columns)
        cols_to_color = index_to_color + cols_to_color + 1

        # Acceder a las celdas del DataFrame original y resaltarlas
        for index in filtered_df_index:
            # + 2, bc df idx start at 0, excel start at 1 and 1 is header
            num_row_to_color = df.index.get_loc(index) + 2
            # range 2, para evitar colorear la primera col 
            for col_num in range(2, cols_to_color):
                cell = excel_sheet.cell(row=num_row_to_color, column=col_num)
                cell.fill = fill

    # Ajustar automáticamente el ancho de las columnas al contenido
    for column in excel_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        excel_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Guardar el archivo Excel nuevamente después de ajustar el ancho de las columnas
    excel_writer.close()


    return nombre_archivo




def insertar_imagen_en_excel(excel_path, imagen_path, nombre_hoja='Sheet1', celda='K1'):
    # Cargar el libro de trabajo de Excel
    libro_excel = openpyxl.load_workbook(excel_path)

    # Obtener la hoja de trabajo (puedes cambiar 'Sheet1' por el nombre de la hoja deseada)
    hoja_trabajo = libro_excel[nombre_hoja]

    # Crear un objeto Image a partir del archivo de imagen
    imagen = Image(imagen_path)

    # Establecer la posición de la celda en la que deseas insertar la imagen
    imagen.anchor = celda

    # Insertar la imagen en la hoja de trabajo
    hoja_trabajo.add_image(imagen)

    # Guardar los cambios en el libro de trabajo
    libro_excel.save(excel_path)


def write_column_to_excel(dataframe, column_name, excel_file_path, start_cell):
    """
    Inserta una columna específica de un DataFrame en un archivo Excel existente
    a partir de una celda específica.

    Parameters:
    - dataframe: DataFrame de pandas
    - column_name: Nombre de la columna que se va a insertar
    - excel_file_path: Ruta del archivo Excel existente
    - start_cell: Celda de inicio para la inserción (por ejemplo, 'B27')

    Returns:
    None
    """
    # Cargar el libro de trabajo existente
    book = load_workbook(excel_file_path)

    # Seleccionar la hoja de trabajo
    sheet = book.active

    # Obtener las coordenadas de la celda de inicio
    start_row = int(start_cell[1:])
    start_col = ord(start_cell[0].upper()) - 65 + 1  # Convertir letra a número de columna

    # Obtener las filas de la columna del DataFrame
    column_data = list(dataframe[column_name])
    column_data = list(dataframe[column_name])[:10]

    # Escribir los datos en las celdas correspondientes
    for idx, value in enumerate(column_data):
        sheet.cell(row=start_row + idx, column=start_col, value=value)

    # Guardar el libro de trabajo actualizado
    book.save(excel_file_path)

    # print(f"Columna '{column_name}' insertada correctamente en la celda {start_cell} del archivo Excel.")


def copy_rows_below(nombre_archivo, inicio_fila, fin_fila):
    # Cargar el libro de trabajo
    libro_trabajo = openpyxl.load_workbook(nombre_archivo)

    # Seleccionar la primera hoja del libro de trabajo
    hoja = libro_trabajo.active

    # Obtener el rango de filas especificado
    rango_a_copiar = hoja[f"A{inicio_fila}:Z{fin_fila}"]


    # Copiar las filas al final de la hoja
    for fila in rango_a_copiar:
        nueva_fila = [copy(celda) for celda in fila]
        hoja.append(nueva_fila)

    # Ajustar la altura de las filas copiadas para que coincida con las originales
    offset = fin_fila - inicio_fila + 1
    for i in range(inicio_fila, fin_fila + 1):
        hoja.row_dimensions[i + offset].height = hoja.row_dimensions[i].height


    # Guardar los cambios en el archivo
    libro_trabajo.save(nombre_archivo)



def copiar_fila_n_veces(nombre_archivo, numero_fila_original, numero_fila_destino, veces_a_copiar):
    # Cargar el libro de trabajo
    libro_trabajo = openpyxl.load_workbook(nombre_archivo)

    # Seleccionar la primera hoja del libro de trabajo
    hoja = libro_trabajo.active

    # Obtener la fila original
    fila_original = hoja[f"A{numero_fila_original}:Z{numero_fila_original}"][0]

    # Copiar la fila con formato y anchura original N veces
    for row_num in range(numero_fila_destino, numero_fila_destino + veces_a_copiar):
        nueva_fila = [copy(celda) for celda in fila_original]
        hoja.append(nueva_fila)

        # Ajustar la altura de la fila copiada para que coincida con la original
        hoja.row_dimensions[row_num].height = hoja.row_dimensions[numero_fila_original].height

    # Guardar los cambios en el archivo
    libro_trabajo.save(nombre_archivo)




def escribir_en_celda(nombre_archivo, nombre_celda, valor):
    # Cargar el archivo Excel
    libro_excel = openpyxl.load_workbook(nombre_archivo)

    # Seleccionar la hoja de trabajo activa
    hoja_activa = libro_excel.active

    # Obtener la celda
    celda = hoja_activa[nombre_celda]

    # Descomentar estas líneas si deseas deshacer la fusión antes de escribir en la celda
    # if celda.merged_cells:
    #     hoja_activa.unmerge_cells(celda_nombre)

    # Escribir el valor en la celda
    celda.value = valor

    # Guardar los cambios en el archivo Excel
    libro_excel.save(nombre_archivo)


def suma_celda_y_entero(nombre_celda, entero):
    # Extraer la parte alfabética del nombre de la celda (por ejemplo, "A" de "A1")
    parte_alfabetica = ''.join(filter(str.isalpha, nombre_celda))

    # Extraer la parte numérica del nombre de la celda (por ejemplo, "1" de "A1")
    parte_numerica = ''.join(filter(str.isdigit, nombre_celda))

    # Sumar el entero a la parte numérica y obtener el nuevo nombre de la celda
    nuevo_numero = int(parte_numerica) + entero
    nuevo_nombre_celda = f"{parte_alfabetica}{nuevo_numero}"

    return nuevo_nombre_celda



def copiar_archivo(origen, destino):
    # Obtener el nombre del archivo original
    file_dst_name = os.path.basename(destino)
    dir_dst_name = os.path.dirname(destino)
    
    # Ruta completa del archivo de destino
    destino_completo = os.path.join(dir_dst_name, file_dst_name)

    # Copiar el archivo de origen al destino
    os.makedirs(dir_dst_name, exist_ok=True)
    shutil.copy2(origen, destino_completo, follow_symlinks=True)